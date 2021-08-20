import requests
import lxml.html
import openpyxl


def parse(url):
    response = requests.get(url)
    tree = lxml.html.document_fromstring(response.text)
    words = tree.xpath('/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()')
    return words


def write_to_xls(number, words, wb):
    page_name = f'Page {number}'
    wb.create_sheet(title=page_name)
    sheet = wb[page_name]
    for word in words:
        cell = sheet.cell(row=words.index(word) + 1, column=1)
        cell.value = word


def main():

    url = 'https://www.allscrabblewords.com/{number}-letter-words/'
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    for i in range(2, 13):
        content = parse(url.format(number=i))
        write_to_xls(i, content, wb)
        # print(content)
        # print(f'{i = }, {content}')

    wb.save('scrabble.xlsx')

def test():
    wb = openpyxl.load_workbook('scrabble.xlsx')
    print(wb.sheetnames)
    sheet = wb['Page 2']
    print('Hello world')
    print('Hello world')
    print('Goodbye World')
 


if __name__ == '__main__':
    test()
